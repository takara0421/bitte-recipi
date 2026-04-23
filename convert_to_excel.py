import re
import pandas as pd
import os
from openpyxl.styles import PatternFill, Font

# Markers used in Markdown to identify sub-recipes
SUB_RECIPE_MARKER = "[サブレシピ]"       # #### [サブレシピ] ... → sub-recipe definition
SUB_RECIPE_REF_MARKER = "[→サブレシピ]"  # ingredient name suffix → reference to a sub-recipe

base_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(base_dir, "transcribed_recipes.md")
output_file = os.path.join(base_dir, "recipes_all.xlsx")


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_markdown_recipes(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    data = []
    current_image_source = ""
    current_recipe_name = ""
    current_section = ""          # #### level — sub-recipe name or section title
    current_group = ""            # ** / < / ( ) inline group markers
    is_sub_recipe_section = False
    current_sub_recipe_name = ""

    in_table = False
    table_headers = []

    for line in lines:
        line = line.strip()

        if not line:
            in_table = False
            continue

        # ── ## Image / chapter header ────────────────────────────────────
        if line.startswith("## "):
            current_image_source = line.replace("## ", "").strip()
            current_recipe_name = ""
            current_section = ""
            current_group = ""
            is_sub_recipe_section = False
            current_sub_recipe_name = ""
            in_table = False
            continue

        # ── ### Recipe name header ────────────────────────────────────────
        if line.startswith("### "):
            current_recipe_name = line.replace("### ", "").strip()
            current_section = ""
            current_group = ""
            is_sub_recipe_section = False
            current_sub_recipe_name = ""
            in_table = False
            continue

        # ── #### Sub-recipe definition  OR  plain section header ──────────
        if line.startswith("#### "):
            section_title = line.replace("#### ", "").strip()
            in_table = False
            current_group = ""

            if SUB_RECIPE_MARKER in section_title:
                is_sub_recipe_section = True
                current_sub_recipe_name = section_title.replace(SUB_RECIPE_MARKER, "").strip()
                current_section = current_sub_recipe_name
            else:
                is_sub_recipe_section = False
                current_section = section_title
            continue

        # ── Table rows ────────────────────────────────────────────────────
        if line.startswith("|"):
            if "---" in line:
                continue

            cells = [c.strip() for c in line.split("|") if c]

            if not in_table:
                table_headers = cells
                in_table = True
            else:
                ingredient_raw = cells[0]

                # Detect sub-recipe reference in the ingredient cell
                is_sub_ref = SUB_RECIPE_REF_MARKER in ingredient_raw
                ingredient = (ingredient_raw
                              .replace(SUB_RECIPE_REF_MARKER, "")
                              .replace("**", "")
                              .strip())

                for i in range(1, len(cells)):
                    if i >= len(table_headers):
                        break
                    variant = table_headers[i]
                    quantity = cells[i]
                    if not quantity or quantity == "-" or quantity.lower() == "nan":
                        continue

                    row_type = (
                        "サブレシピ定義" if is_sub_recipe_section
                        else "サブレシピ参照" if is_sub_ref
                        else "レシピ"
                    )
                    data.append({
                        "Recipe": current_recipe_name,
                        "Section": current_section or "Main",
                        "Group": current_group,
                        "Variant": variant,
                        "Ingredient": ingredient,
                        "Quantity": quantity,
                        "Row Type": row_type,
                        "Sub-Recipe Name": current_sub_recipe_name if is_sub_recipe_section
                                           else (ingredient if is_sub_ref else ""),
                        "Source Image": current_image_source,
                    })
            continue

        # ── Bullet list items ─────────────────────────────────────────────
        if line.startswith("* ") or line.startswith("- "):
            content = line.lstrip("*- ").strip()

            # Inline group markers (bold / angle-bracket / paren headings)
            if (content.startswith("**") or content.startswith("<")
                    or (content.startswith("(") and content.endswith(")"))):
                current_group = content
                continue

            is_sub_ref = SUB_RECIPE_REF_MARKER in content
            content_clean = content.replace(SUB_RECIPE_REF_MARKER, "").strip()

            ingredient = ""
            quantity = ""

            if ":" in content_clean:
                parts = content_clean.split(":", 1)
                ingredient = parts[0].strip()
                quantity = parts[1].strip()
            else:
                match = re.search(r'\s(\d+(?:[\.~]\d+)?[a-zA-Z%コ個]+)$', content_clean)
                if match:
                    quantity = match.group(1)
                    ingredient = content_clean[:match.start()].strip()
                else:
                    ingredient = content_clean

            row_type = (
                "サブレシピ定義" if is_sub_recipe_section
                else "サブレシピ参照" if is_sub_ref
                else "レシピ"
            )
            data.append({
                "Recipe": current_recipe_name,
                "Section": current_section or "Main",
                "Group": current_group,
                "Variant": current_group or "Standard",
                "Ingredient": ingredient,
                "Quantity": quantity,
                "Row Type": row_type,
                "Sub-Recipe Name": current_sub_recipe_name if is_sub_recipe_section
                                   else (ingredient if is_sub_ref else ""),
                "Source Image": current_image_source,
            })
            continue

        # ── Inline group / section headers (non-bullet) ───────────────────
        if (line.startswith("**") or line.startswith("<")
                or (line.startswith("(") and line.endswith(")"))):
            current_group = line
            continue

    return data


# ---------------------------------------------------------------------------
# Cleaning helpers
# ---------------------------------------------------------------------------

def clean_recipe_name(name):
    """Remove parenthetical English translations from recipe names."""
    name = re.sub(r'\s*\(.*?\)', '', name)
    name = re.sub(r'\s*<.*?>', '', name)
    return name.strip()


def clean_quantity(quantity_str):
    """Extract a numeric gram value from a quantity string."""
    if not quantity_str:
        return ""

    # Prefer bracketed weight, e.g. "2.5個 (160g)" → "160g"
    if '(' in str(quantity_str) and ')' in str(quantity_str):
        m = re.search(r'\((.*?)\)', str(quantity_str))
        if m:
            inner = m.group(1)
            if 'g' in inner or 'ml' in inner:
                quantity_str = inner

    m = re.search(r'(\d+(?:\.\d+)?)', str(quantity_str))
    if m:
        return float(m.group(1))
    return 0


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def _auto_col_width(worksheet, df, max_width=40):
    for idx, col in enumerate(df.columns):
        col_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
        if idx < 26:
            worksheet.column_dimensions[chr(65 + idx)].width = min(col_len, max_width)


def _apply_row_highlighting(worksheet, df, col_order):
    """Highlight sub-recipe reference rows in yellow."""
    if "Row Type" not in col_order:
        return

    ref_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ref_font = Font(bold=True, color="B8860B")
    rt_col = col_order.index("Row Type") + 1
    ing_col = col_order.index("Ingredient") + 1 if "Ingredient" in col_order else None

    for row_idx in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row_idx, column=rt_col).value == "サブレシピ参照":
            for c in range(1, len(col_order) + 1):
                worksheet.cell(row=row_idx, column=c).fill = ref_fill
            if ing_col:
                worksheet.cell(row=row_idx, column=ing_col).font = ref_font


def create_excel(data, output_path):
    if not data:
        print("No data found to write.")
        return

    df = pd.DataFrame(data)

    if not df.empty:
        df['Recipe'] = df['Recipe'].apply(clean_recipe_name)
        df['Quantity'] = df['Quantity'].apply(clean_quantity)

    # ── Split into main recipes and sub-recipe definitions ────────────────
    df_sub  = df[df['Row Type'] == 'サブレシピ定義'].copy()
    df_main = df[df['Row Type'] != 'サブレシピ定義'].copy()

    cols_main = ["Recipe", "Section", "Group", "Variant",
                 "Ingredient", "Quantity", "Row Type", "Sub-Recipe Name", "Source Image"]
    cols_sub  = ["Recipe", "Sub-Recipe Name", "Variant",
                 "Ingredient", "Quantity", "Source Image"]

    cols_main = [c for c in cols_main if c in df_main.columns]
    cols_sub  = [c for c in cols_sub  if c in df_sub.columns]

    df_main = df_main[cols_main]
    df_sub  = df_sub[cols_sub]

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1 – All recipes (with sub-recipe references highlighted)
        df_main.to_excel(writer, index=False, sheet_name="全レシピ")
        _auto_col_width(writer.sheets["全レシピ"], df_main)
        _apply_row_highlighting(writer.sheets["全レシピ"], df_main, cols_main)

        # Sheet 2 – Sub-recipe definitions
        if not df_sub.empty:
            df_sub.to_excel(writer, index=False, sheet_name="サブレシピ")
            _auto_col_width(writer.sheets["サブレシピ"], df_sub)

    print(f"Excel file created at: {output_path}")
    print(f"  全レシピ sheet : {len(df_main)} rows")
    if not df_sub.empty:
        print(f"  サブレシピ sheet: {len(df_sub)} rows  "
              f"({df_sub['Sub-Recipe Name'].nunique()} sub-recipe(s))")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if not os.path.exists(input_file):
        print(f"Error: Input file not found at {input_file}")
    else:
        parsed_data = parse_markdown_recipes(input_file)
        if not parsed_data:
            print("No data found to write.")
        else:
            create_excel(parsed_data, output_file)
